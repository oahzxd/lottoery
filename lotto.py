#!/usr/bin/python
#coding=utf-8

#author   : bmhzxd
#version  : 1.0
#datetime : 2015.7.27

import os
import re
import sys
import time
import string
import urllib2
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles.colors import RED

fontRed = Font(color=RED, bold=True)
fontBold = Font(bold=True)

redBalls = [[0 for col in range(6)] for row in range(500)]
blueBalls = [0 for row in range(500)]
omitTable = [[-1 for col in range(34)] for row in range(500)]

arrSeral = []
arrDate = []
startSeral = 0

"获取网络数据"
"sNo: 最近期数"
"sNo=0: 为输入截止日期方法查询"
def getData(sNo):
    global arrSeral, arrDate
    url = 'http://baidu.lecai.com/lottery/draw/list/50?type='
    if (sNo != 0):
        url += 'latest&num=' + str(sNo)
    else:
        startDate = ''
        endDate = ''
        url += 'range_date&start=' + startDate + '&end=' + endDate
        
    page = urllib2.urlopen(url)
    soup = BeautifulSoup(page)
    #print soup.prettify()
    
    row = 0
    col = 0
    for red in soup.find_all('td', class_='redBalls'):
        for ball in red.find_all('em'):
            redBalls[row][col] = int(ball.text)
            col = col + 1
            #print ball.text,
       #print
        col = 0
        row = row + 1
        
    row = 0
    for blue in soup.find_all('td', class_='blueBalls'):
        for ball in blue.find_all('em'):
            blueBalls[row] = int(ball.text)
            row = row + 1
            #print ball.text,
        #print
    
    tbody = soup.find_all('tbody')[0]
    for seral in tbody.find_all(href=re.compile('draw')):
        date = seral.find_parents('td')[0].find_next_siblings('td')[0]
        arrSeral.append(seral.text)
        arrDate.append(date.text)
        #print seral.text
        #print date.text
        
    #seral = tbody.find_all(href=re.compile('draw'))[0]
    #gDate = seral.find_parents('td')[0].find_next_siblings('td')[0].text
    #gSeral = seral.text
    setOmitTable()

    
def setOmitTable():
    for i in range(1, 34): # 1-33号码
        omit = 0
        for j in range(99, -1, -1):
            for k in range(0, 6):
                if redBalls[j][k] == i:
                    omit = 0
            omitTable[j][i] = omit
            omit = omit + 1

def applyXlsFont(ws):
    rowNum = ws.max_row
    ws['A%d' % rowNum].font = fontBold
    ws['B%d' % rowNum].font = fontBold
    ws['C%d' % rowNum].font = fontBold
    ws['D%d' % rowNum].font = fontBold
    ws['E%d' % rowNum].font = fontBold
    ws['F%d' % rowNum].font = fontBold
    ws['G%d' % rowNum].font = fontBold
    ws['H%d' % rowNum].font = fontBold
    ws['I%d' % rowNum].font = fontBold
    ws['J%d' % rowNum].font = fontBold
    ws['K%d' % rowNum].font = fontBold
    ws['L%d' % rowNum].font = fontRed
    ws['M%d' % rowNum].font = fontBold
    ws['N%d' % rowNum].font = fontBold
    ws['O%d' % rowNum].font = fontBold
    ws['P%d' % rowNum].font = fontBold
    ws['Q%d' % rowNum].font = fontBold
    ws['R%d' % rowNum].font = fontBold
    ws['S%d' % rowNum].font = fontBold
    ws['T%d' % rowNum].font = fontBold
    ws['U%d' % rowNum].font = fontBold
    
            
"1.奇偶数"
def fileOddOrEven(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    applyXlsFont(ws)
    oddOrEven(ws, 10)
    oddOrEven(ws, 5)
    oddOrEven(ws, 1)

    
def oddOrEven(ws, sNo):
    odd = 0
    even = 0
    for i in range(startSeral, startSeral + sNo):
        for j in range(0, 6):
            if (redBalls[i][j] % 2 == 0):
                even = even + 1
            else:
                odd = odd + 1

    if (sNo != 1):
        if (even > odd):
            ws.append(['近' + str(sNo) + '期有', '奇=' + str(odd) + '，偶=' + str(even), '偶+' + str(even - odd)])
            applyXlsFont(ws)
        elif (even < odd):
            ws.append(['近' + str(sNo) + '期有', '奇=' + str(odd) + '，偶=' + str(even), '奇+' + str(odd - even)])
            applyXlsFont(ws)
        else:
            ws.append(['近' + str(sNo) + '期有', '奇=' + str(odd) + '，偶=' + str(even), '奇=偶'])
            applyXlsFont(ws)
    else:
        ws.append(['本期奇偶比例:', '奇 : 偶 = %d : %d' % (odd, even)])
        applyXlsFont(ws)


"2.大小数"
def fileBigOrSmall(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    applyXlsFont(ws)
    bigOrSmall(ws, 10)
    bigOrSmall(ws, 5)
    bigOrSmall(ws, 1)

    
def bigOrSmall(ws, sNo):
    big = 0
    small = 0
    for i in range(startSeral, startSeral + sNo):
        for j in range(0, 6):
            if (redBalls[i][j] <= 16):
                small = small + 1
            else:
                big = big + 1

    if (sNo != 1):
        if (small > big):
            ws.append(['近' + str(sNo) + '期有', '大=' + str(big) + '，小=' + str(small), '小+' + str(small - big)])
            applyXlsFont(ws)
        elif (small < big):
            ws.append(['近' + str(sNo) + '期有', '大=' + str(big) + '，小=' + str(small), '大+' + str(big - small)])
            applyXlsFont(ws)
        else:
            ws.append(['近' + str(sNo) + '期有', '大=' + str(big) + '，小=' + str(small), '大=小'])
            applyXlsFont(ws)
    else:
        ws.append(['本期大小比例:', '大 : 小 = %d : %d' % (big, small)])
        applyXlsFont(ws)

        
"3.和值偏差"
def sumOffset(ws):
    global arrDate, arrSeral
    sum = 0
    for i in range(0, 6):
        sum += redBalls[startSeral][i]
    cells = ['' for row in range(21)]
    cells[0] = arrDate[startSeral]
    cells[1] = arrSeral[startSeral]
    cells[2] = str(sum)
    cells[11] = '+'

    place = sum / 10
    # sum: cells[place + 1]  100: 11
    if place > 10:
        for i in range(12, place + 1 + 1):
            cells[i] = '+'
    elif place < 10:
        for i in range(place + 1, 11):
            cells[i] = '+'
    ws.append(cells)
    applyXlsFont(ws)

    
"4.遗漏"
def fileOmit(ws):
    global arrDate, arrSeral
    cells = ['  ' for row in range(7)]
    cells[0] = arrDate[startSeral]
    cells[1] = arrSeral[startSeral]
    for i in range(0, 6):
        cells[2] += '  %02d' % (redBalls[startSeral][i])

    o = omit(startSeral)
            
    omitNum = 0 # 小于10的个数
    omitSum = 0 # 总和
    for i in range(0, 6):
        if o[i] < 10:
            omitNum = omitNum + 1
        cells[3] += '  %02d' % (o[i])
        omitSum = omitSum + o[i]

    cells[4] = str(omitNum)
    cells[5] = str(omitSum)
    cells[6] = str(float('%.1f' % (omitSum / 6.0)))
    ws.append(cells)
    applyXlsFont(ws)

    rowNum = ws.max_row
    numCell = ws['E%d' % rowNum]
    if numCell.value == '6':
        numCell.font = fontRed

        
def omit(start):
    o = [-1 for col in range(6)] #遗漏次数
    for i in range(0, 6):
        for j in range(start + 1, 99):
            for k in range(0, 6):
                if redBalls[j][k] == redBalls[start][i]:
                    o[i] = j - start - 1
                    break
            if o[i] != -1:
                break
    return o


"5.区间"
def ballRange(ws):
    global arrDate, arrSeral
    cells = ['' for row in range(9)]
    cells[0] = arrDate[startSeral]
    cells[1] = arrSeral[startSeral]

    bRange = [0 for col in range(7)]
    for i in range(0, 6):
        if redBalls[startSeral][i] >= 1 and redBalls[startSeral][i] <= 5:
            bRange[0] = bRange[0] + 1
        elif redBalls[startSeral][i] >= 6 and redBalls[startSeral][i] <= 10:
            bRange[1] = bRange[1] + 1
        elif redBalls[startSeral][i] >= 11 and redBalls[startSeral][i] <= 15:
            bRange[2] = bRange[2] + 1
        elif redBalls[startSeral][i] >= 16 and redBalls[startSeral][i] <= 20:
            bRange[3] = bRange[3] + 1
        elif redBalls[startSeral][i] >= 21 and redBalls[startSeral][i] <= 25:
            bRange[4] = bRange[4] + 1
        elif redBalls[startSeral][i] >= 26 and redBalls[startSeral][i] <= 30:
            bRange[5] = bRange[5] + 1
        elif redBalls[startSeral][i] >= 31 and redBalls[startSeral][i] <= 33:
            bRange[6] = bRange[6] + 1
            
    for i in range(0, 7):
        if bRange[i] == 0:
            cells[i+2] = '-'
        else:
            cells[i+2] = str(bRange[i])
    ws.append(cells)
    applyXlsFont(ws)

    
"6.百分比表, 热号，冷号，温号"
def hotOrCold(ws):
    global arrDate, arrSeral
    nums = [0 for row in range(33)]
    cells = ['' for row in range(5)]
    cells[0] = arrDate[startSeral]
    cells[1] = arrSeral[startSeral]
    for i in range(startSeral, startSeral + 5):
        for j in range(0, 6):
            if nums[redBalls[i][j] - 1] == 0:
                nums[redBalls[i][j] - 1] = 1 #近1-5期出现
    for i in range(startSeral + 5, startSeral + 10):
        for j in range(0, 6):
            if nums[redBalls[i][j] - 1] == 0:
                nums[redBalls[i][j] - 1] = 2 #近6-10期出现
            elif nums[redBalls[i][j] -1] == 1:
                nums[redBalls[i][j] - 1] = 3 #两个区间都出现
    hot = '  '
    cold = '  '
    warm = '  '
    hNum = 0
    cNum = 0
    wNum = 0
    for i in range(0, 33):
        if nums[i] == 0:
            cold += '  %02d' % (i + 1)
            cNum = cNum + 1
            if (cNum >= 10):
                cNum = 0
                #cold += '\n'
        elif nums[i] == 3:
            hot += '  %02d' % (i + 1)
            hNum = hNum + 1
            if (hNum >= 10):
                hNum = 0
                #hot += '\n'
        else:
            warm += '  %02d' % (i + 1)
            wNum = wNum + 1
            if (wNum >= 10):
                wNum = 0
                #warm += '\n'

    cells[2] = hot
    cells[3] = cold
    cells[4] = warm
    ws.append(cells)
    applyXlsFont(ws)

    
"7.遗漏数字偏差表"
def omitDict(start, o):
    out = False
    for i in range(0, 6):
        for j in range(start + 1, 99):
            for k in range(0, 6):
                if redBalls[j][k] == redBalls[start][i]:
                    o[redBalls[start][i]] = j -start - 1
                    out = True
                    break
            if out:
                out = False
                break
    return o

def omitOffset(ws, d):
    global arrDate, arrSeral

    redStr = ''
    for i in range(0, 6):
        redStr += '%02d ' % redBalls[startSeral][i]

    ws.append(['过去' + str(d) + '期', '', '本期中奖号码: %s' % redStr])
    applyXlsFont(ws)
    ws.append(['遗漏次数', '符合个数', '符合数字'])
    applyXlsFont(ws)

    o = {}
    for i in range(startSeral + d - 1, startSeral - 1, -1):
        omitDict(i, o)
    arr = [[] for i in range(6)]
    cnt = {}
    for i in range(0, 6):
        for (k,v) in o.items():
            if v == i:
                arr[i].append(k)
        cnt[i] = len(arr[i])
    dictCnt = sorted(cnt.iteritems(), key=lambda d:d[1], reverse = False)
    
    for i in range(0, 6):
        diStr = '  '
        for j in range(1, 34):
            if omitTable[startSeral][j] == dictCnt[i][0]:
                diStr += '  %02d' % (j)
        ws.append([dictCnt[i][0], dictCnt[i][1], diStr])
        applyXlsFont(ws)

        
def colorOmitOffset(ws, d):
    global arrDate, arrSeral, startSeral
    if startSeral >= 49:
        return

    startSeral = startSeral + 1
    redStr = ''
    rowNum = ws.max_row - 15
    o = {}
    for i in range(startSeral + d - 1, startSeral - 1, -1):
        omitDict(i, o)
    arr = [[] for i in range(6)]
    cnt = {}
    for i in range(0, 6):
        for (k,v) in o.items():
            if v == i:
                arr[i].append(k)
        cnt[i] = len(arr[i])
    dictCnt = sorted(cnt.iteritems(), key=lambda d:d[1], reverse = False)
    
    for i in range(0, 6):
        redStr = ''
        for j in range(1, 34):
            if omitTable[startSeral][j] == dictCnt[i][0]:
                for cnt in range(0, 6):
                    if j == redBalls[startSeral - 1][cnt]:
                        redStr += '%02d  ' % j
        
        ws['F%d' % rowNum].font = fontRed
        ws['F%d' % rowNum].value = redStr
        rowNum = rowNum + 1
        
    startSeral = startSeral -1

def fileOmitOffset(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    applyXlsFont(ws)
    omitOffset(ws, 5)
    colorOmitOffset(ws, 5)
        
    #omitOffset(ws, 6)
    #omitOffset(ws, 7)
    #omitOffset(ws, 8)
    #omitOffset(ws, 9)
    #omitOffset(ws, 10)

    
"(中)博彩趋势逆转"
def countOmit(ball, start):
    for i in range(start, 99):
        for j in range(0, 6):
            if redBalls[i][j] == ball:
                return i - start

            
def trendReverse(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    applyXlsFont(ws)

    parn = ['' for row in range(6)]
    find = False
    num = 1

    # 筛选最近4期号码, 只查询上期中奖数字
    for i in range(0, 6):
        for j in range(startSeral + 4, startSeral, -1):
            for k in range(0, 6):
                if redBalls[j][k] == redBalls[startSeral][i]:
                    parn[i] += 'x'
                    find = True
                    num = 1
                    break
            if find == False:
                parn[i] += str(num)
                num = num + 1
            find = False
        parn[i] += 'x'
        num = 1

    find = False
    for i in range(0, 6):
        if parn[i] == '123xx':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om + 3 >= 17:
                ws.append(['博彩逆转：', '数字：' + str(redBalls[startSeral][i]), '遗漏了' + str(om + 3) + '次', ('模式是1..%d ' % (om + 3)) + parn[i][3:5]])
                applyXlsFont(ws)
                find = True
        elif parn[i] == '12x1x':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om + 2 >= 17:
                ws.append(['博彩逆转：', '数字：' + str(redBalls[startSeral][i]), '遗漏了' + str(om + 2) + '次', ('模式是1..%d ' % (om + 2)) + parn[i][2:5]])
                applyXlsFont(ws)
                find = True
        elif parn[i] == '1x12x':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om + 1 >= 17:
                ws.append(['博彩逆转：', '数字：' + str(redBalls[startSeral][i]), '遗漏了' + str(om + 1) + '次', ('模式是1..%d ' % (om + 1)) + parn[i][1:5]])
                applyXlsFont(ws)
                find = True
        elif parn[i] == 'x123x':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om >= 17:
                ws.append(['博彩逆转：', '数字：' + str(redBalls[startSeral][i]), '遗漏了' + str(om) + '次', ('模式是1..%d ' % (om)) + parn[i][:]])
                applyXlsFont(ws)
                find = True
    
"层叠"
def pile(ws):
    # 只查询中奖的号码
    for num in range(0, 6):
        ser = startSeral + 1
        ball = redBalls[startSeral][num]
        omit = omitTable[ser][ball]
        omit2 = 0
        omit3 = 0
        parn = []
        if omit >= 15:
            # x
            omit2 = omitTable[ser + omit + 1][ball]
            if omit2 == 3:
                # x123x
                parn.append('x123x 1..%d x' % omit)
                ser = ser + 4
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 2:
                # x12x
                parn.append('x12x 1..%d x' % omit)
                ser = ser + 3
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 1:
                # x1x
                parn.append('x1x 1..%d x' % omit)
                ser = ser + 2
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 0:
                # xx
                parn.append('xx 1..%d x' % omit)
                ser = ser + 1
                omit2 = omitTable[ser + omit + 1][ball]
                i = 0
                while omit2 == 0:
                    # xxx..
                    i = i + 1
                    ser = ser + 1
                    omit2 = omitTable[ser + omit + 1][ball]
                    parn.append('x')
            else:
                parn.append('x 1..%d x' % omit)
                
            if omit2 >= 8 and omit2 <= 15:
                # x
                omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                if omit3 == 3:
                    # x123x
                    ser = ser + 4
                    parn.append('x123x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 2:
                    # x12x
                    ser = ser + 3
                    parn.append('x12x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3== 1:
                    # x1x
                    ser = ser + 2
                    parn.append('x1x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 0:
                    # xx
                    ser = ser + 1
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                    parn.append('xx 1..%d ' % omit2)
                    i = 0
                    while omit3 == 0:
                        # xxx..
                        i = i + 1
                        ser = ser + 1
                        omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                        parn.append('x')
                else:
                    parn.append('x 1..%d ' % omit2)

                if omit3 >= 3 and omit3 <= 8:
                    parn.append('x 1..%d ' % omit3)
                    parn.reverse()
                    ws.append(['层叠：', '数字：%02d' % (ball) , '模式：%s' % ''.join(parn)])
                    applyXlsFont(ws)

    
"反向层叠"
def rePile(ws):
    # 只查询中奖的号码
    for num in range(0, 6):
        ser = startSeral + 1
        ball = redBalls[startSeral][num]
        omit = omitTable[ser][ball]
        omit2 = 0
        omit3 = 0
        parn = []
        if omit <=8 and omit >= 3:
            # x
            omit2 = omitTable[ser + omit + 1][ball]
            if omit2 == 3:
                # x123x
                parn.append('x123x 1..%d x' % omit)
                ser = ser + 4
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 2:
                # x12x
                parn.append('x12x 1..%d x' % omit)
                ser = ser + 3
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 1:
                # x1x
                parn.append('x1x 1..%d x' % omit)
                ser = ser + 2
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 0:
                # xx
                parn.append('xx 1..%d x' % omit)
                ser = ser + 1
                omit2 = omitTable[ser + omit + 1][ball]
                i = 0
                while omit2 == 0:
                    # xxx..
                    i = i + 1
                    ser = ser + 1
                    omit2 = omitTable[ser + omit + 1][ball]
                    parn.append('x')
            else:
                parn.append('x 1..%d x' % omit)
                
            if omit2 >= 8 and omit2 <= 15:
                # x
                omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                if omit3 == 3:
                    # x123x
                    ser = ser + 4
                    parn.append('x123x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 2:
                    # x12x
                    ser = ser + 3
                    parn.append('x12x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3== 1:
                    # x1x
                    ser = ser + 2
                    parn.append('x1x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 0:
                    # xx
                    ser = ser + 1
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                    parn.append('xx 1..%d ' % omit2)
                    i = 0
                    while omit3 == 0:
                        # xxx..
                        i = i + 1
                        ser = ser + 1
                        omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                        parn.append('x')
                else:
                    parn.append('x 1..%d ' % omit2)

                if omit3 >= 15:
                    parn.append('x 1..%d ' % omit3)
                    parn.reverse()
                    ws.append(['反向层叠：', '数字：%02d' % (ball) , '模式：%s' % ''.join(parn)])
                    applyXlsFont(ws)

        
"n底"
def nBottom(ws):
    # 检验所有数字，预测提示
    for num in range(1, 34):
        ser = startSeral
        ball = num
        cnt = 1
        parn = []
        tRange = -1
        if omitTable[ser][ball] == 0:
            continue
        
        while True:
            omit = omitTable[ser][ball]
            i = 0
            while omit == 0:
                i = i + 1
                ser = ser + 1
                omit = omitTable[ser][ball]
                
            omit2 = omitTable[ser + omit + 1][ball]
            j = 0
            while omit2 == 0:
                j = j + 1
                ser = ser + 1
                omit2 = omitTable[ser + omit + 1][ball]
                
            if tRange == -1:
                tRange = omit

            if (omit >= 2 and omit2 >= 2) and (omit == tRange or omit == tRange + 1 or omit == tRange - 1) and (omit2 == tRange or omit2 == tRange + 1 or omit2 == tRange - 1):
                if cnt == 1:
                    for cnt in range(omit, 0, -1):
                        parn.append(str(cnt))

                while i != 0:
                    parn.append('x')
                    i = i - 1
                    
                parn.append('x')
                cnt = cnt + 1
                ser = ser + omit + 1
                
                while j != 0:
                    parn.append('x')
                    j = j - 1

                for k in range(omit2, 0, -1):
                    parn.append(str(k))


            else:
                if cnt != 1:
                    parn.append('x')
                    parn.reverse()
                    ws.append(['%d倍底：' % cnt, '数字：%d' % ball, '模式：%s' % ''.join(parn)])
                    applyXlsFont(ws)
                    cnt = 1
                break


"旗式排列"
def flagRange(ws):
    # 检验所有数字，预测提示
    # tips: 利用数组翻转
    for num in range(1, 34):
        ser = startSeral
        ball = num
        parn = []
        isOk = False
        #isOk2 = False
        if omitTable[ser][ball] == 0:
            continue
        isInit = True
        omit = 0
        cnt = 0
        while True:
            omit = omitTable[ser][ball]
            if isInit == True:
                if omit != 5 and omit != 6:
                    isOk = True
                    break
                
            if omit == 5 or omit == 6:
                ser = ser + omit + 1
                if isInit == False:
                    parn.append('x')    
                for j in range(omit, 0, -1):
                    parn.append(str(j))
                if ser > 99:
                    break
                isInit = False
            elif omit < 3:
                # if omit == 3:
                #     #x123x
                #     parn.append('x123x')
                #     ser = ser + omit + 1
                if omit == 2:
                    #x12x
                    parn.append('x12x')
                    ser = ser + omit + 1
                elif omit == 1:
                    #x1x
                    parn.append('x1x')
                    ser = ser + omit + 1
                elif omit == 0:
                    #xx
                    parn.append('x')
                    while omit == 0:
                        ser = ser + 1
                        omit = omitTable[ser][ball]
                        parn.append('x')
                isInit = True
            else:
                break
            
            cnt = cnt + 1

        if cnt >= 2 and isOk == True:
            parn.reverse()
            ws.append(['旗式排列：', '数字：%d' % ball, '模式：%s' % ''.join(parn)])
            applyXlsFont(ws)


"热门中奖数字 -- 废弃"
def hotNum(ws):
    for i in range(0, 6):
        ser = startSeral + 1
        ball = redBalls[startSeral][i]
        parn = []
        isOk = False
        isOk2 = False
        while True:
            omit = omitTable[ser][ball]
            if omit >= 6:
                isOk = True
                ser = ser + omit + 1
                parn.append('x')
                for j in range(omit, 0, -1):
                    parn.append(str(j))
                if ser > 99:
                    break

            else:
                break

        if isOk:
            while True:
                if omit == 2 or omit == 1:
                    parn.append('x')
                    for j in range(omit, 0, -1):
                        parn.append(str(j))
                    isOk2 = True
                    ser = ser + omit + 1
                    omit = omitTable[ser][ball]
                    if ser > 99:
                        break
                else:
                    if isOk2:
                        parn.append('x')
                        parn.reverse()
                        ws.append(['热门中奖数字：', '数字：%d' % ball, '模式：%s' % ''.join(parn)])
                        applyXlsFont(ws)
                        ws.append([''])
                    break

                
"创建新表"
def createXlsx():
    wb = Workbook()
    
    ws0 = wb.create_sheet(0)
    ws0.title = '文件信息'
    ws0['A1'] = '最新更新期数'
    
    ws1 = wb.create_sheet(1)
    ws1.title = '奇偶数表'

    ws2 = wb.create_sheet(2)
    ws2.title = '大小数表'

    ws3 = wb.create_sheet(3)
    ws3.title = '和值偏差表'
    ws3.append(['日期', '期数', '和值', '020', '030', '040', '050', '060', '070', '080', '090', '100', '110', '120', '130', '140', '150', '160', '170', '180', '190'])
    applyXlsFont(ws3)

    ws4 = wb.create_sheet(4)
    ws4.title = '区间表'
    ws4.append(['日期', '期数', '1-5', '6-10', '11-15', '16-20', '21-25', '26-30', '31-33'])
    applyXlsFont(ws4)

    ws5 = wb.create_sheet(5)
    ws5.title = '百分比表'
    ws5.append(['日期', '期数', '热号', '冷号', '温号'])
    applyXlsFont(ws5)

    ws6 = wb.create_sheet(6)
    ws6.title = '遗漏表'
    ws6.append(['日期', '期数', '中奖号码', '遗漏情况', '遗漏少于10次的个数', '总计', '平均'])
    applyXlsFont(ws6)

    ws7 = wb.create_sheet(7)
    ws7.title = '遗漏数字偏差表'

    ws8 = wb.create_sheet(8)
    ws8.title = '中期表'
    
    return wb

        
"完整性检查"
def checkComplete(ws):
    global arrDate, arrSeral, startSeral
    v = ws['B1'].value
    if v == arrSeral[0]:
        print 'no need to update.'
        exit()
    elif v == arrSeral[1]:
        startSeral = 0
        print 'update the newest.'
        return True
    else:
        for i in range(2, 100):
            if v == arrSeral[i]:
                startSeral = i - 1 # 从下一期开始更新
                print 'update from %s' % arrSeral[startSeral]
                break
        return False

    
def addInfo(ws):
    ws['B1'] = arrSeral[0]

    
def countBall():
    fileOddOrEven(wb['奇偶数表'])
    fileBigOrSmall(wb['大小数表'])
    sumOffset(wb['和值偏差表'])
    ballRange(wb['区间表'])
    hotOrCold(wb['百分比表'])
    fileOmit(wb['遗漏表'])
    fileOmitOffset(wb['遗漏数字偏差表'])
    
    trendReverse(wb['中期表'])
    pile(wb['中期表'])
    rePile(wb['中期表'])
    nBottom(wb['中期表'])
    flagRange(wb['中期表'])
    #hotNum(wb['中期表'])
    
if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding('utf-8')
    exist = os.path.isfile('../data.xlsx')
    getData(100) #近100期
    if exist:
        print 'load exsit file.'
        wb = load_workbook('../data.xlsx')
        if checkComplete(wb['文件信息']):
            countBall()
        else:
            while (startSeral != -1):
                countBall()
                startSeral = startSeral - 1
    else:
        print 'create new file.'
        wb = createXlsx()
        startSeral = 49
        while (startSeral != -1):
            countBall()
            startSeral = startSeral - 1

    addInfo(wb['文件信息'])
    wb.save('../data.xlsx')
